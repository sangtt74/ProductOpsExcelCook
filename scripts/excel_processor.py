# scripts/excel_processor.py
import openpyxl
import os
import pandas as pd # Có thể cần nếu bạn muốn trả về DataFrame

# Import các hàm từ code_generator nếu cần dùng chúng
from .code_generator import load_existing_codes, generate_random_code, get_unique_filename

def process_excel_for_codes(uploaded_excel_file, directory_to_check_codes, output_dir, progress_callback_excel=None):
    """
    Processes an Excel file to generate codes based on prefixes and quantities.
    Returns a list of paths to generated files.
    """
    generated_file_paths = []
    rows_processed = 0

    try:
        workbook = openpyxl.load_workbook(uploaded_excel_file)
        sheet = workbook.active
    except Exception as e:
        raise ValueError(f"Error loading XLSX file: {e}. Make sure it's a valid XLSX file.")

    # Determine start row (skip header if present)
    start_row = 1
    if sheet.max_row >= 1 and len(sheet[1]) >= 2:
        first_row_values = [cell.value for cell in sheet[1]]
        if isinstance(first_row_values[0], str) and isinstance(first_row_values[1], str) and \
           "prefix" in first_row_values[0].lower() and "quantity" in first_row_values[1].lower():
            start_row = 2

    total_rows_to_process = sheet.max_row - start_row + 1
    if total_rows_to_process <= 0:
        raise ValueError("Excel file contains no data rows to process or only headers.")

    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row), start=start_row):
        if progress_callback_excel:
            progress_callback_excel(min(1.0, (row_idx - start_row + 1) / total_rows_to_process), f"Processing row: {row_idx} / {sheet.max_row}")

        if len(row) < 2:
            print(f"Warning: Skipping row {row_idx} as it does not have enough columns (expected 2).")
            continue

        prefix_cell_value = row[0].value
        num_codes_cell_value = row[1].value

        prefix = str(prefix_cell_value).strip().upper() if prefix_cell_value is not None else ""
        try:
            num_codes = int(num_codes_cell_value) if num_codes_cell_value is not None else 0
        except (ValueError, TypeError):
            num_codes = 0

        if not (3 <= len(prefix) <= 8):
            print(f"Error in row {row_idx}: Prefix '{prefix}' is not between 3 and 8 characters long. Skipping this row.")
            continue

        if num_codes <= 0:
            print(f"Error in row {row_idx}: Number of codes '{num_codes}' must be greater than 0. Skipping this row.")
            continue

        # Use load_existing_codes from the same package
        existing_codes_set = load_existing_codes(directory_to_check_codes, prefix)

        try:
            # Pass a dummy callback or a real Streamlit callback for generate_random_code
            # Here we use a lambda that just prints, as its own progress is for its internal loop
            codes_to_write = generate_random_code(prefix, num_codes, existing_codes_set, 
                                                  progress_callback=lambda p, s: print(f"  Internal progress for {prefix}: {s}"))

            if codes_to_write:
                output_file_path = get_unique_filename(prefix, output_dir)

                with open(output_file_path, mode="w", newline="", encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(["code"])
                    writer.writerows(codes_to_write)

                generated_file_paths.append(output_file_path)
                rows_processed += 1
            else:
                print(f"No new unique codes could be generated for prefix '{prefix}'. Skipping CSV creation for this row.")

        except ValueError as e:
            print(f"Error in row {row_idx} for prefix '{prefix}': {e}. Skipping this row.")
        except Exception as e:
            print(f"An unexpected error occurred while processing row {row_idx} for prefix '{prefix}': {e}. Skipping this row.")

    return generated_file_paths, rows_processed

# Bạn có thể thêm các hàm khác liên quan đến việc xử lý Excel ở đây